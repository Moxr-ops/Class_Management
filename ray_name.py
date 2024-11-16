import openpyxl
import sys

class Ray_name:

    def __init__(self):
        self.name_list = []
        self.missing_name_list = []

    def filter_text(self, text): #检测文本中是否包含名字
        for item in self.name_list:
            if item not in text:
                self.missing_name_list.append(item)
    
    def get_names_from_excel(self, excel_path):# 从花名册中获得同学姓名    
        excel_path = excel_path.replace('"', '')
        try:
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
            name_list = []

            for column in sheet.iter_cols(min_row=1, max_row=sheet.max_row, values_only=True):
                for cell in column:
                    if cell is not None and "姓名" in str(cell):
                        name_list.extend([cell for cell in column if cell is not None])
                        break
            if name_list == []:
                print("未能找到包含“姓名”单元格的列，请重新选择花名册")
                return False
            self.name_list = name_list
            print("\n检索到的名字有：\n")
            for name in self.name_list:
                print(f"{name}\n")
            print(f"共{len(self.name_list)}个")
            return True
        except Exception as e:
            print(f"文件打开失败，请确保输入的是Excel文件的路径。错误信息：{e}")
            return False

                

