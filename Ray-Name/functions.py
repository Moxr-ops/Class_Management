import openpyxl

class Ray_name:

    def __init__(self):
        self.name_list = []

    def filter_text(self, text): #检测文本中是否包含名字
        filtered_list = []
        for item in self.name_list:
            if item not in text:
                filtered_list.append(item)

        return filtered_list
    
    def get_names_from_excel(self, excel_path): #从花名册中获得同学姓名
        excel_path = excel_path.replace('"', '')
        print(excel_path)
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        name_list = []

        for column in sheet.iter_cols(min_row=1, max_row=sheet.max_row, values_only=True):
            for cell in column:
                if cell is not None and "姓名" in str(cell):
                    name_list.extend([cell for cell in column if cell is not None])
                    break
        self.name_list = name_list

                

